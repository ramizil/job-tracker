"""Export applications to CSV or Excel."""
from __future__ import annotations

import csv
import io
from datetime import date

from .tracker import list_applications

COLUMNS = [
    "id", "starred", "company", "title", "location", "status", "match_score",
    "ai_fit_score", "ai_fit_level", "source", "url", "salary", "contact",
    "resume_version", "date_found", "date_applied", "rejection_stage",
    "rejection_reason", "rejection_note", "rejection_date", "ai_verdict",
    "notes",
]

# Header labels shown in Excel (friendlier than raw column keys).
_HEADER_LABELS = {
    "id": "#",
    "starred": "★",
    "match_score": "Match %",
    "ai_fit_score": "AI %",
    "ai_fit_level": "AI Level",
    "date_found": "Found",
    "date_applied": "Applied",
    "rejection_stage": "Rejection Stage",
    "rejection_reason": "Rejection Reason",
    "rejection_note": "Rejection Note",
    "rejection_date": "Rejected",
    "ai_verdict": "AI Verdict",
    "resume_version": "Resume",
}

# Same solid colors as the HTML status pills / AI badges (style.css).
_STATUS_COLORS = {
    "saved": "F59E0B",
    "applied": "2563EB",
    "reapplied": "0D9488",
    "screening": "CA8A04",
    "interview": "9333EA",
    "offer": "16A34A",
    "accepted": "16A34A",
    "rejected": "DC2626",
    "withdrawn": "64748B",
    "ghosted": "64748B",
}
_FIT_COLORS = {
    "yes": "16A34A",
    "maybe": "CA8A04",
    "no": "DC2626",
}

_DATE_COLS = {"date_found", "date_applied", "rejection_date"}
_WIDE_COLS = {"title", "notes", "url", "ai_verdict", "rejection_note"}


def _header_label(col: str) -> str:
    if col in _HEADER_LABELS:
        return _HEADER_LABELS[col]
    return col.replace("_", " ").title()


def _cell_value(row, col: str):
    if col not in row.keys():
        return ""
    val = row[col]
    if val is None:
        return ""
    if col == "starred":
        return "★" if val else ""
    if col in _DATE_COLS:
        s = str(val).strip()
        return s[:10] if s else ""
    return val


def to_csv() -> str:
    rows = list_applications(order_by="id DESC")
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({c: _cell_value(r, c) for c in COLUMNS})
    return buf.getvalue()


def to_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = list_applications(order_by="id DESC")
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    white_bold = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    body_font = Font(name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    zebra = PatternFill("solid", fgColor="F8FAFC")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Header row
    ws.append([_header_label(c) for c in COLUMNS])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[1].height = 22

    col_index = {c: i + 1 for i, c in enumerate(COLUMNS)}
    status_col = col_index["status"]
    fit_score_col = col_index["ai_fit_score"]
    fit_level_col = col_index["ai_fit_level"]
    starred_col = col_index["starred"]
    url_col = col_index["url"]
    id_col = col_index["id"]

    for r_i, r in enumerate(rows, start=2):
        values = [_cell_value(r, c) for c in COLUMNS]
        ws.append(values)
        stripe = (r_i % 2 == 0)

        status = (str(r["status"] or "")).strip().lower() if "status" in r.keys() else ""
        fit_level = (
            str(r["ai_fit_level"] or "").strip().lower()
            if "ai_fit_level" in r.keys() else ""
        )
        status_hex = _STATUS_COLORS.get(status)
        fit_hex = _FIT_COLORS.get(fit_level)

        for c_i, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_i, column=c_i)
            cell.font = body_font
            cell.border = thin

            if col in ("status", "ai_fit_score", "ai_fit_level", "starred",
                       "id", "match_score"):
                cell.alignment = center
            elif col in _WIDE_COLS and col != "url":
                cell.alignment = wrap
            else:
                cell.alignment = left

            if stripe and c_i not in (status_col, fit_score_col, fit_level_col):
                cell.fill = zebra

        # Status pill colors
        if status_hex:
            sc = ws.cell(row=r_i, column=status_col)
            sc.fill = PatternFill("solid", fgColor=status_hex)
            sc.font = white_bold
            sc.alignment = center

        # AI % + level badge colors
        if fit_hex:
            for ci in (fit_score_col, fit_level_col):
                fc = ws.cell(row=r_i, column=ci)
                fc.fill = PatternFill("solid", fgColor=fit_hex)
                fc.font = white_bold
                fc.alignment = center

        # Star highlight
        star_cell = ws.cell(row=r_i, column=starred_col)
        if star_cell.value:
            star_cell.font = Font(bold=True, color="CA8A04", name="Calibri", size=12)
            star_cell.alignment = center

        # Clickable job URL
        url_val = ws.cell(row=r_i, column=url_col).value
        if url_val and str(url_val).startswith(("http://", "https://")):
            url_cell = ws.cell(row=r_i, column=url_col)
            url_cell.hyperlink = str(url_val)
            url_cell.font = Font(name="Calibri", size=11, color="2563EB",
                                 underline="single")

        # Numeric id / scores stay numbers for filter & sort
        id_cell = ws.cell(row=r_i, column=id_col)
        if id_cell.value not in ("", None):
            try:
                id_cell.value = int(id_cell.value)
            except (TypeError, ValueError):
                pass
        for score_col_name in ("match_score", "ai_fit_score"):
            scell = ws.cell(row=r_i, column=col_index[score_col_name])
            if scell.value not in ("", None):
                try:
                    scell.value = int(scell.value)
                except (TypeError, ValueError):
                    pass

    # Column widths
    for i, c in enumerate(COLUMNS, 1):
        if c == "id":
            width = 6
        elif c == "starred":
            width = 4
        elif c in ("status", "ai_fit_level", "source"):
            width = 12
        elif c in ("match_score", "ai_fit_score"):
            width = 9
        elif c in ("company", "location"):
            width = 22
        elif c in _DATE_COLS:
            width = 12
        elif c in ("title", "ai_verdict"):
            width = 36
        elif c in ("notes", "rejection_note"):
            width = 40
        elif c == "url":
            width = 28
        else:
            width = 16
        ws.column_dimensions[get_column_letter(i)].width = width

    # AutoFilter + freeze header — filter dropdowns on every column
    last_row = max(1, len(rows) + 1)
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    # Print / view niceties
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:1"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def download_name() -> str:
    """Suggested Excel filename with today's date."""
    return f"applications_{date.today().isoformat()}.xlsx"
